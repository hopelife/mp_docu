# [파이썬 크롤러 만들기](https://jione-e.tistory.com/28)
# [Source Code for Package lxml.html](https://lxml.de/api/lxml.html-pysrc.html)

import os, sys
import json
import time
import re
import openpyxl
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Side, Border

import lxml.html as ht
import lxml.etree as et

##------------------------------------------------------------
## User 모듈
##------------------------------------------------------------
sys.path.append(os.path.join(os.path.dirname(__file__), '../_public')) ## Note: 현재 디렉토리 기준 상대 경로 설정
from utils_basic import (_create_folder, _read_file, _file_to_json, _to_lists, _fn)
from utils_xlsx import (_xsheet, _write_xlsx, _set_width, _set_height, _apply_style_theme)

def _normalize_spaces(s):
    """연결된 공백을 하나의 공백으로 변경
    """
    return re.sub(r'\s+', ' ', s).strip()


def _remove_punc(s):
    """ 특수문자 제거

    Args:
        s (str): 입력 문자열
    """
    punc = '[!"#$%&\'()*+,-./:;<=>?[\]^_`{|}~“”·]'
    return re.sub(punc, '', s)


def _dicts_from_html(path, xpath):
    root = ht.fromstring(_read_file(path))
    for key in keys:
        locals()[key] = ''
    
    ## xpath: elements xpath
    options = {
        'key1': {
            'xpath': '', ##
            'attr': 'text'
        }
    }
    pass


if __name__ == '__main__':
    root_path = '../../../__references/_python/sats'
    xlsx_path = f'{root_path}/pattern_candles.xlsx'
    title = 'wikipedia'

    sheet = _xsheet(path=xlsx_path, title=title)
    wb = sheet['wb']
    ws = sheet['ws']

    # C:\Dev\docMoon\__references\_python\sats\scrap\wikipedia_org
    ## NOTE: after scraping
    path = f'{root_path}/scrap/wikipedia_org/Candlestick_pattern.html'
    root = ht.fromstring(_read_file(path))

    names = []
    meanings = []
    images = []

    for tr in root.xpath('.//*[@id="mw-content-text"]/div[1]//tr'):
        tds = tr.xpath('./td')

        if len(tds) != 4:  # NOTE: td가 4개가 아닌 행이면
            continue
        # elif  tds[1].text == None:  # NOTE: 비어있는 행이면
        #     continue
        print(f"tds len: {len(tds)}")
        for i, td in enumerate(tds):
            try:
                # image = images.append(td.xpath('.//img')[0]
                if i%2 == 0:  # NOTE: image 포함 열
                    images.append(td.xpath('.//img')[0].attrib['src'])
                else:
                    names.append(td.xpath('.//b')[0].text_content())
                    meanings.append(td.text_content())
            except:
                print("의미없는 행")

    zips = zip(images, names, meanings)
    candles = []
    for image, name, meaning in zips:
        candles.append({'image': image, 'name': name.replace(' Candle', '').replace(' ', '_'), 'meaning': meaning})


    ## NOTE: cell header 입력
    for i, col in enumerate(candles[0].keys()):
        # NOTE: name
        ws.cell(row=1, column=i+1, value=col)
    
    ws.cell(row=1, column=4, value="condition")

    ## NOTE: cell contents 입력
    for i, candle in enumerate(candles):
        # col = openpyxl.utils.get_column_letter(i + 2)
        # NOTE: image
        _path = f"{root_path}/scrap/wikipedia_org/{candle['image']}"
        try:
            img = openpyxl.drawing.image.Image(_path)
            img.anchor = f'A{i+2}'
            ws.add_image(img)
        except:
            print("이미지 삽입 실패: {_path}")

        # NOTE: name
        ws.cell(row=i+2, column=2, value=candle['name'])

        # NOTE: meaning
        ws.cell(row=i+2, column=3, value=candle['meaning'])


    ## 셀 너비/높이 지정    
    row_max = len(candles)
    # col_max = len(list(candles[0].keys()))

    widths = dict(A=13, B=25, C=80, D=40)
    _set_width(ws, widths)
    heights = {1: 20, (2, row_max+1): 75}
    print(f"row_max+1: {row_max+1}")
    _set_height(ws, heights)
    

    # wb = _write_xlsx(sheet, candles, theme='default')
    _apply_style_theme(sheet, theme='default', count_row=row_max+1, count_col=4, header=0, first=(0, 0))
    ws.cell(row=1, column=4).style = 'header_style'  ## 데이터 없는 셀 스타일 지정
    
    _create_folder('/'.join(xlsx_path.split('/')[:-1]))
    print(f"path: {xlsx_path}")
    wb.save(xlsx_path)

    